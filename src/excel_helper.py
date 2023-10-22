import openpyxl as xl
from openpyxl.styles import Border, Side

from helper import reverse_map

tl = Border(left=Side(style='thick'), right=Side(style='thin'), top=Side(style='thick'), bottom=Side(style='thin'))
tr = Border(left=Side(style='thin'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thin'))
br = Border(left=Side(style='thin'), right=Side(style='thick'), top=Side(style='thin'), bottom=Side(style='thick'))
bl = Border(left=Side(style='thick'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thick'))

ml = Border(left=Side(style='thick'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
mr = Border(left=Side(style='thin'), right=Side(style='thick'), top=Side(style='thin'), bottom=Side(style='thin'))
mt = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thick'), bottom=Side(style='thin'))
mb = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thick'))
cc = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

border_delta_map = {
    (0, 0): tl,
    (1, 0): ml,
    (2, 0): bl,
    (0, 1): mt,
    (0, 2): tr,
    (1, 2): mr,
    (2, 1): mb,
    (2, 2): br,
    (1, 1): cc
}


def apply_borders(file_path):
    wb = xl.load_workbook(file_path)
    sheet = wb['Sheet1']
    for r in range(3):
        for c in range(3):
            for rd in range(3):
                for cd in range(3):
                    sheet.cell(r * 3 + 2 + rd, c * 3 + 2 + cd).border = border_delta_map[(rd, cd)]

    wb.save(file_path)


def reformat_excel(file_path):
    wb = xl.load_workbook(file_path)
    ws = wb.active
    ws.sheet_view.showGridLines = False
    sheet = wb['Sheet1']

    for r in range(1, 11):
        sheet.cell(r, 1, value='').border = Border()

    for c in range(1, 11):
        sheet.cell(1, c, value='').border = Border()

    for r in range(2, 11):
        for c in range(2, 11):
            sheet.cell(r, c).font = xl.styles.Font(size=20, bold=True, color='000000')
            sheet.cell(r, c).alignment = xl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)

    for x in range(1, 11):
        sheet.row_dimensions[x].height = 35
        sheet.column_dimensions[xl.utils.get_column_letter(x)].width = 7

    wb.save(file_path)


def print_candidates_in_excel(output_path, candidates):
    reversed_map = reverse_map(candidates)
    wb = xl.load_workbook(output_path)
    sheet = wb['Sheet1']

    for (r, c), digits in reversed_map.items():
        ds = ", ".join(str(d) for d in digits)
        # sheet.cell(r+2, c+2).value = ds
        sheet.cell(r + 2, c + 2, value=ds).font = xl.styles.Font(size=9, bold=False, color='FF0000')
        sheet.cell(r + 2, c + 2).alignment = xl.styles.Alignment(horizontal='left', vertical='top', wrap_text=True)
        sheet.cell(r + 2, c + 2).fill = xl.styles.PatternFill(patternType='solid', fgColor='FFFF00')

    wb.save(output_path)
